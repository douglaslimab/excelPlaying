from openpyxl import load_workbook, open, Workbook
from os import mkdir

book = Workbook()
files_names = []

mkdir('./Sheets')

for i in range(10):
    files_names.append('05-0800-0' + str(i) + '.Xlsx')


for name in files_names:
    mkdir('./Sheets/' + str(name[:-4]))
    book['Sheet']['A1'] = 'File Name'
    book['Sheet']['B1'] = name[:-5]
    book.save('./Sheets/' + str(name[:-5])+'/'+name)

for bk in files_names:
    sav = load_workbook('./Sheets/' + str(bk[:-5])+'/'+bk)
    print(sav['Sheet']['B1'].value)
    print(sav.sheetnames)